"""
Rotas para Ordens de Compra
"""
from fastapi import APIRouter, Depends, HTTPException, Request, Cookie
from fastapi.responses import HTMLResponse, RedirectResponse, Response
from sqlalchemy.orm import Session
from typing import Optional
import logging
import io
import requests
from PIL import Image as PILImage
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.worksheet.hyperlink import Hyperlink
from datetime import datetime
import tempfile
import os

from app.config.database import get_db
from app.controllers.ordem_compra_controller import OrdemCompraController
from app.controllers.auth_controller import AuthController
from app.views.template_renderer import render_template
from app.models.saas_models import OrdemCompra, OrdemCompraItem, Fornecedor

logger = logging.getLogger(__name__)

ordem_compra_router = APIRouter()
ordem_compra_controller = OrdemCompraController()
auth_controller = AuthController()

def get_company_id_from_user(user_data: dict) -> int:
    """Extrair company_id dos dados do usuário"""
    return user_data.get("company_id")

# Rotas HTML
@ordem_compra_router.get("/ordem-compra", response_class=HTMLResponse)
async def ordem_compra_page(
    request: Request,
    session_token: Optional[str] = Cookie(None),
    db: Session = Depends(get_db)
):
    """Página de ordens de compra"""
    if not session_token:
        return RedirectResponse(url="/auth/login", status_code=302)
    
    result = auth_controller.get_user_by_session(session_token, db)
    if result.get("error"):
        return RedirectResponse(url="/auth/login", status_code=302)
    
    user_data = result["user"]
    
    return render_template("ordem_compra.html", user=user_data)

@ordem_compra_router.get("/ordem-compra/nova", response_class=HTMLResponse)
async def nova_ordem_compra_page(
    request: Request,
    session_token: Optional[str] = Cookie(None),
    db: Session = Depends(get_db)
):
    """Página de nova ordem de compra"""
    if not session_token:
        return RedirectResponse(url="/auth/login", status_code=302)
    
    result = auth_controller.get_user_by_session(session_token, db)
    if result.get("error"):
        return RedirectResponse(url="/auth/login", status_code=302)
    
    user_data = result["user"]
    
    return render_template("nova_ordem_compra.html", user=user_data)

@ordem_compra_router.get("/ordem-compra/editar/{ordem_id}", response_class=HTMLResponse)
async def editar_ordem_compra_page(
    ordem_id: int,
    request: Request,
    session_token: Optional[str] = Cookie(None),
    db: Session = Depends(get_db)
):
    """Página de edição de ordem de compra"""
    if not session_token:
        return RedirectResponse(url="/auth/login", status_code=302)
    
    result = auth_controller.get_user_by_session(session_token, db)
    if result.get("error"):
        return RedirectResponse(url="/auth/login", status_code=302)
    
    user_data = result["user"]
    company_id = get_company_id_from_user(user_data)
    
    # Buscar dados da ordem
    ordem = db.query(OrdemCompra).filter(
        OrdemCompra.id == ordem_id,
        OrdemCompra.company_id == company_id
    ).first()
    
    if not ordem:
        return RedirectResponse(url="/ordem-compra", status_code=302)
    
    # Buscar itens da ordem
    itens = db.query(OrdemCompraItem).filter(
        OrdemCompraItem.ordem_compra_id == ordem_id
    ).all()
    
    # Preparar dados da ordem
    ordem_data = {
        "id": ordem.id,
        "numero_ordem": ordem.numero_ordem,
        "data_ordem": ordem.data_ordem.isoformat() if ordem.data_ordem else None,
        "data_entrega_prevista": ordem.data_entrega_prevista.isoformat() if ordem.data_entrega_prevista else None,
        "status": ordem.status,
        "valor_total": float(ordem.valor_total or 0),
        "desconto": float(ordem.desconto or 0),
        "valor_final": float(ordem.valor_final or 0),
        "moeda": ordem.moeda,
        "cotacao_moeda": float(ordem.cotacao_moeda or 1.0),
        "tipo_ordem": ordem.tipo_ordem,
        "comissao_agente": float(ordem.comissao_agente or 0),
        "percentual_comissao": float(ordem.percentual_comissao or 0),
        "valor_transporte": float(ordem.valor_transporte or 0),
        "percentual_importacao": float(ordem.percentual_importacao or 0),
        "taxas_adicionais": float(ordem.taxas_adicionais or 0),
        "valor_impostos": float(ordem.valor_impostos or 0),
        "fornecedor_id": ordem.fornecedor_id,
        "fornecedor_nome": ordem.fornecedor.nome if ordem.fornecedor else None,
        "transportadora_id": ordem.transportadora_id,
        "transportadora_nome": ordem.transportadora.nome if ordem.transportadora else None,
        "observacoes": ordem.observacoes,
        "itens": [
            {
                "id": item.id,
                "produto_nome": item.produto_nome,
                "produto_descricao": item.produto_descricao,
                "produto_codigo": item.produto_codigo,
                "produto_imagem": item.produto_imagem,
                "descricao_fornecedor": item.descricao_fornecedor,
                "quantidade": float(item.quantidade),
                "valor_unitario": float(item.valor_unitario),
                "valor_total": float(item.valor_total),
                "url": item.url,
                "observacoes": item.observacoes
            }
            for item in itens
        ]
    }
    
    # Log para debug
    logger.info(f"Dados da ordem carregados: {ordem_data}")
    logger.info(f"Número de itens: {len(ordem_data['itens'])}")
    
    return render_template("nova_ordem_compra.html", user=user_data, ordem_data=ordem_data, is_editing=True)

# Rotas API
@ordem_compra_router.get("/api/ordem-compra")
async def get_ordens_compra(
    status: Optional[str] = None,
    search: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    session_token: Optional[str] = Cookie(None),
    db: Session = Depends(get_db)
):
    """API para listar ordens de compra"""
    if not session_token:
        raise HTTPException(status_code=401, detail="Token de sessão necessário")
    
    result = auth_controller.get_user_by_session(session_token, db)
    if result.get("error"):
        raise HTTPException(status_code=401, detail="Sessão inválida ou expirada")
    
    user_data = result["user"]
    company_id = get_company_id_from_user(user_data)
    
    ordens = ordem_compra_controller.get_ordens_compra(company_id, db, status, search, date_from, date_to)
    
    return {
        "success": True,
        "ordens": ordens,
        "total": len(ordens)
    }

@ordem_compra_router.get("/api/ordem-compra/{ordem_id}")
async def get_ordem_compra(
    ordem_id: int,
    session_token: Optional[str] = Cookie(None),
    db: Session = Depends(get_db)
):
    """API para buscar ordem de compra por ID"""
    if not session_token:
        raise HTTPException(status_code=401, detail="Token de sessão necessário")
    
    result = auth_controller.get_user_by_session(session_token, db)
    if result.get("error"):
        raise HTTPException(status_code=401, detail="Sessão inválida ou expirada")
    
    user_data = result["user"]
    company_id = get_company_id_from_user(user_data)
    
    ordem = ordem_compra_controller.get_ordem_compra_by_id(ordem_id, company_id, db)
    
    if not ordem:
        raise HTTPException(status_code=404, detail="Ordem de compra não encontrada")
    
    return {
        "success": True,
        "ordem": ordem
    }

@ordem_compra_router.post("/api/ordem-compra")
async def create_ordem_compra(
    request: Request,
    session_token: Optional[str] = Cookie(None),
    db: Session = Depends(get_db)
):
    """API para criar ordem de compra"""
    if not session_token:
        raise HTTPException(status_code=401, detail="Token de sessão necessário")
    
    result = auth_controller.get_user_by_session(session_token, db)
    if result.get("error"):
        raise HTTPException(status_code=401, detail="Sessão inválida ou expirada")
    
    user_data = result["user"]
    company_id = get_company_id_from_user(user_data)
    
    # Obter dados do corpo da requisição
    body = await request.json()
    
    # Validações básicas
    if not body.get('numero_ordem') and not body.get('itens'):
        raise HTTPException(status_code=400, detail="Número da ordem ou itens são obrigatórios")
    
    result = ordem_compra_controller.create_ordem_compra(body, company_id, db)
    
    if not result.get("success"):
        raise HTTPException(status_code=400, detail=result.get("error", "Erro ao criar ordem de compra"))
    
    return result

@ordem_compra_router.put("/api/ordem-compra/{ordem_id}")
async def update_ordem_compra(
    ordem_id: int,
    request: Request,
    session_token: Optional[str] = Cookie(None),
    db: Session = Depends(get_db)
):
    """API para atualizar ordem de compra"""
    if not session_token:
        raise HTTPException(status_code=401, detail="Token de sessão necessário")
    
    result = auth_controller.get_user_by_session(session_token, db)
    if result.get("error"):
        raise HTTPException(status_code=401, detail="Sessão inválida ou expirada")
    
    user_data = result["user"]
    company_id = get_company_id_from_user(user_data)
    
    # Obter dados do corpo da requisição
    body = await request.json()
    
    result = ordem_compra_controller.update_ordem_compra(ordem_id, body, company_id, db)
    
    if not result.get("success"):
        raise HTTPException(status_code=400, detail=result.get("error", "Erro ao atualizar ordem de compra"))
    
    return result

@ordem_compra_router.delete("/api/ordem-compra/{ordem_id}")
async def delete_ordem_compra(
    ordem_id: int,
    session_token: Optional[str] = Cookie(None),
    db: Session = Depends(get_db)
):
    """API para deletar ordem de compra"""
    if not session_token:
        raise HTTPException(status_code=401, detail="Token de sessão necessário")
    
    result = auth_controller.get_user_by_session(session_token, db)
    if result.get("error"):
        raise HTTPException(status_code=401, detail="Sessão inválida ou expirada")
    
    user_data = result["user"]
    company_id = get_company_id_from_user(user_data)
    
    result = ordem_compra_controller.delete_ordem_compra(ordem_id, company_id, db)
    
    if not result.get("success"):
        raise HTTPException(status_code=400, detail=result.get("error", "Erro ao deletar ordem de compra"))
    
    return result

@ordem_compra_router.patch("/api/ordem-compra/{ordem_id}/status")
async def update_status_ordem(
    ordem_id: int,
    request: Request,
    session_token: Optional[str] = Cookie(None),
    db: Session = Depends(get_db)
):
    """API para alterar status da ordem de compra"""
    if not session_token:
        raise HTTPException(status_code=401, detail="Token de sessão necessário")
    
    result = auth_controller.get_user_by_session(session_token, db)
    if result.get("error"):
        raise HTTPException(status_code=401, detail="Sessão inválida ou expirada")
    
    user_data = result["user"]
    company_id = get_company_id_from_user(user_data)
    
    # Obter dados do corpo da requisição
    body = await request.json()
    status = body.get("status")
    
    if not status:
        raise HTTPException(status_code=400, detail="Status é obrigatório")
    
    result = ordem_compra_controller.update_status_ordem(ordem_id, status, company_id, db)
    
    if not result.get("success"):
        raise HTTPException(status_code=400, detail=result.get("error", "Erro ao alterar status da ordem"))
    
    return result

# Rotas para gerenciar links externos
@ordem_compra_router.get("/api/ordem-compra/{ordem_id}/links")
async def get_ordem_compra_links(
    ordem_id: int,
    session_token: Optional[str] = Cookie(None),
    db: Session = Depends(get_db)
):
    """API para buscar links externos de uma ordem de compra"""
    if not session_token:
        raise HTTPException(status_code=401, detail="Token de sessão necessário")
    
    result = auth_controller.get_user_by_session(session_token, db)
    if result.get("error"):
        raise HTTPException(status_code=401, detail="Sessão inválida ou expirada")
    
    user_data = result["user"]
    company_id = get_company_id_from_user(user_data)
    
    links = ordem_compra_controller.get_ordem_compra_links(ordem_id, company_id, db)
    
    return {"success": True, "links": links}

@ordem_compra_router.post("/api/ordem-compra/{ordem_id}/links")
async def add_ordem_compra_link(
    ordem_id: int,
    request: Request,
    session_token: Optional[str] = Cookie(None),
    db: Session = Depends(get_db)
):
    """API para adicionar link externo a uma ordem de compra"""
    if not session_token:
        raise HTTPException(status_code=401, detail="Token de sessão necessário")
    
    result = auth_controller.get_user_by_session(session_token, db)
    if result.get("error"):
        raise HTTPException(status_code=401, detail="Sessão inválida ou expirada")
    
    user_data = result["user"]
    company_id = get_company_id_from_user(user_data)
    
    # Obter dados do corpo da requisição
    body = await request.json()
    
    result = ordem_compra_controller.add_ordem_compra_link(ordem_id, company_id, body, db)
    
    if not result.get("success"):
        raise HTTPException(status_code=400, detail=result.get("error", "Erro ao adicionar link"))
    
    return result

@ordem_compra_router.put("/api/ordem-compra/links/{link_id}")
async def update_ordem_compra_link(
    link_id: int,
    request: Request,
    session_token: Optional[str] = Cookie(None),
    db: Session = Depends(get_db)
):
    """API para atualizar link externo"""
    if not session_token:
        raise HTTPException(status_code=401, detail="Token de sessão necessário")
    
    result = auth_controller.get_user_by_session(session_token, db)
    if result.get("error"):
        raise HTTPException(status_code=401, detail="Sessão inválida ou expirada")
    
    user_data = result["user"]
    company_id = get_company_id_from_user(user_data)
    
    # Obter dados do corpo da requisição
    body = await request.json()
    
    result = ordem_compra_controller.update_ordem_compra_link(link_id, company_id, body, db)
    
    if not result.get("success"):
        raise HTTPException(status_code=400, detail=result.get("error", "Erro ao atualizar link"))
    
    return result

@ordem_compra_router.delete("/api/ordem-compra/links/{link_id}")
async def delete_ordem_compra_link(
    link_id: int,
    session_token: Optional[str] = Cookie(None),
    db: Session = Depends(get_db)
):
    """API para deletar link externo"""
    if not session_token:
        raise HTTPException(status_code=401, detail="Token de sessão necessário")
    
    result = auth_controller.get_user_by_session(session_token, db)
    if result.get("error"):
        raise HTTPException(status_code=401, detail="Sessão inválida ou expirada")
    
    user_data = result["user"]
    company_id = get_company_id_from_user(user_data)
    
    result = ordem_compra_controller.delete_ordem_compra_link(link_id, company_id, db)
    
    if not result.get("success"):
        raise HTTPException(status_code=400, detail=result.get("error", "Erro ao deletar link"))
    
    return result

@ordem_compra_router.put("/api/ordem-compra/{ordem_id}/status")
async def update_ordem_compra_status(
    ordem_id: int,
    request: Request,
    session_token: Optional[str] = Cookie(None),
    db: Session = Depends(get_db)
):
    """API para alterar status de uma ordem de compra"""
    try:
        # Verificar autenticação
        if not session_token:
            raise HTTPException(status_code=401, detail="Token de sessão não fornecido")
        
        # Buscar sessão
        from app.models.saas_models import UserSession
        session = db.query(UserSession).filter(UserSession.session_token == session_token).first()
        if not session:
            raise HTTPException(status_code=401, detail="Sessão inválida")
        
        # Buscar usuário para obter company_id
        from app.models.saas_models import User
        user = db.query(User).filter(User.id == session.user_id).first()
        if not user:
            raise HTTPException(status_code=401, detail="Usuário não encontrado")
        
        company_id = user.company_id
        
        # Buscar dados do corpo da requisição
        body = await request.json()
        new_status = body.get('status')
        
        if not new_status:
            raise HTTPException(status_code=400, detail="Status não fornecido")
        
        # Atualizar status
        result = ordem_compra_controller.update_ordem_compra_status(ordem_id, company_id, new_status, db)
        
        if not result.get("success"):
            raise HTTPException(status_code=400, detail=result.get("error", "Erro ao alterar status"))
        
        return result
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erro na API de alteração de status: {str(e)}")
        raise HTTPException(status_code=500, detail="Erro interno do servidor")

@ordem_compra_router.get("/api/ordem-compra/export/{ordem_id}")
async def export_ordem_compra(
    ordem_id: int,
    session_token: Optional[str] = Cookie(None),
    db: Session = Depends(get_db)
):
    """API para exportar ordem de compra em Excel"""
    if not session_token:
        raise HTTPException(status_code=401, detail="Token de sessão necessário")
    
    result = auth_controller.get_user_by_session(session_token, db)
    if result.get("error"):
        raise HTTPException(status_code=401, detail="Sessão inválida ou expirada")
    
    user_data = result["user"]
    company_id = get_company_id_from_user(user_data)
    
    # Buscar dados completos da ordem
    ordem = ordem_compra_controller.get_ordem_compra_by_id(ordem_id, company_id, db)
    if not ordem:
        raise HTTPException(status_code=404, detail="Ordem de compra não encontrada")
    
    # Buscar dados da empresa
    from app.models.saas_models import Company
    company = db.query(Company).filter(Company.id == company_id).first()
    
    # Buscar dados do fornecedor
    fornecedor = None
    if ordem.get('fornecedor_id'):
        fornecedor = db.query(Fornecedor).filter(Fornecedor.id == ordem['fornecedor_id']).first()
    
    # Criar workbook Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Ordem de Compra"
    
    # Estilos
    title_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    normal_font = Font(name='Arial', size=10)
    bold_font = Font(name='Arial', size=10, bold=True)
    
    title_fill = PatternFill(start_color='2F4F4F', end_color='2F4F4F', fill_type='solid')
    header_fill = PatternFill(start_color='4682B4', end_color='4682B4', fill_type='solid')
    light_fill = PatternFill(start_color='F0F8FF', end_color='F0F8FF', fill_type='solid')
    
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')
    right_alignment = Alignment(horizontal='right', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título principal
    ws.merge_cells('A1:F1')
    ws['A1'] = f"ORDEM DE COMPRA - {ordem.get('numero_ordem', '')}"
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = center_alignment
    ws.row_dimensions[1].height = 30
    
    # Dados da Empresa e Fornecedor lado a lado
    row = 3
    
    # Cabeçalho DADOS DA EMPRESA (lado esquerdo)
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'] = "DADOS DA EMPRESA"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].alignment = left_alignment
    ws.row_dimensions[row].height = 25
    
    # Cabeçalho DADOS DO FORNECEDOR (lado direito)
    ws.merge_cells(f'D{row}:F{row}')
    ws[f'D{row}'] = "DADOS DO FORNECEDOR"
    ws[f'D{row}'].font = header_font
    ws[f'D{row}'].fill = header_fill
    ws[f'D{row}'].alignment = left_alignment
    
    # Dados da Empresa (lado esquerdo)
    if company:
        row += 1
        empresa_data = [
            ['Nome:', company.name or ''],
            ['CNPJ:', company.cnpj or ''],
            ['Endereço:', f"{company.endereco or ''} {company.numero or ''}"],
            ['Cidade:', f"{company.cidade or ''} - {company.estado or ''}"],
            ['CEP:', company.cep or ''],
            ['País:', company.pais or 'Brasil']
        ]
        
        for i, (label, value) in enumerate(empresa_data):
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = bold_font
            ws[f'A{row}'].fill = light_fill
            ws[f'A{row}'].border = thin_border
            
            ws[f'B{row}'] = value
            ws[f'B{row}'].font = normal_font
            ws[f'B{row}'].border = thin_border
            ws.merge_cells(f'B{row}:C{row}')
            row += 1
    
    # Dados do Fornecedor (lado direito)
    if fornecedor:
        fornecedor_data = [
            ['Nome:', fornecedor.nome or ''],
            ['Nome Fantasia:', fornecedor.nome_fantasia or ''],
            ['CNPJ:', fornecedor.cnpj or ''],
            ['Endereço:', f"{fornecedor.endereco or ''} {fornecedor.numero or ''}"],
            ['Cidade:', f"{fornecedor.cidade or ''} - {fornecedor.estado or ''}"],
            ['CEP:', fornecedor.cep or ''],
            ['País:', fornecedor.pais or ''],
            ['Contato:', fornecedor.contato_nome or ''],
            ['Telefone:', fornecedor.telefone or ''],
            ['Email:', fornecedor.email or '']
        ]
        
        # Resetar row para começar do lado direito
        row = 4
        for i, (label, value) in enumerate(fornecedor_data):
            ws[f'D{row}'] = label
            ws[f'D{row}'].font = bold_font
            ws[f'D{row}'].fill = light_fill
            ws[f'D{row}'].border = thin_border
            
            ws[f'E{row}'] = value
            ws[f'E{row}'].font = normal_font
            ws[f'E{row}'].border = thin_border
            ws.merge_cells(f'E{row}:F{row}')
            row += 1
    
    # Ajustar row para a próxima seção (pegar o maior entre empresa e fornecedor)
    if company and fornecedor:
        row = max(len(empresa_data) + 4, len(fornecedor_data) + 4)
    elif company:
        row = len(empresa_data) + 4
    elif fornecedor:
        row = len(fornecedor_data) + 4
    else:
        row = 4
    
    # Dados da Ordem
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "DADOS DA ORDEM"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].alignment = left_alignment
    ws.row_dimensions[row].height = 25
    
    row += 1
    ordem_data = [
        ['Número da Ordem:', ordem.get('numero_ordem', '')],
        ['Data:', ordem.get('data_ordem', '')],
        ['Status:', ordem.get('status', '')],
        ['Moeda:', ordem.get('moeda', 'BRL')],
        ['Cotação:', f"{ordem.get('cotacao_moeda', 1):.4f}"],
        ['Valor Total:', f"{ordem.get('valor_final', 0):.2f}"],
        ['Data de Entrega:', ordem.get('data_entrega_prevista', '')]
    ]
    
    for i, (label, value) in enumerate(ordem_data):
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = bold_font
        ws[f'A{row}'].fill = light_fill
        ws[f'A{row}'].border = thin_border
        
        ws[f'B{row}'] = value
        ws[f'B{row}'].font = normal_font
        ws[f'B{row}'].border = thin_border
        ws.merge_cells(f'B{row}:F{row}')
        row += 1
    
    # Itens da Ordem
    if ordem.get('itens'):
        row += 1
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = "ITENS DA ORDEM"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws[f'A{row}'].alignment = left_alignment
        ws.row_dimensions[row].height = 25
        
        # Cabeçalho da tabela de itens
        row += 1
        headers = ['Produto', 'Quantidade', 'Valor Unit.', 'Total', 'URL', 'Imagem']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border
        
        # Dados dos itens
        for item in ordem['itens']:
            row += 1
            
            # Coluna Produto com todas as informações
            produto_nome = item.get('produto_nome', '')
            produto_sku = item.get('produto_codigo', '')
            descricao_fornecedor = item.get('descricao_fornecedor', '')
            
            # Montar texto completo do produto
            produto_completo = produto_nome
            if produto_sku:
                produto_completo += f"\nSKU: {produto_sku}"
            if descricao_fornecedor:
                produto_completo += f"\nDescrição: {descricao_fornecedor}"
            
            ws[f'A{row}'] = produto_completo
            ws[f'B{row}'] = item.get('quantidade', 0)
            ws[f'C{row}'] = f"{item.get('valor_unitario', 0):.2f}"
            ws[f'D{row}'] = f"{item.get('valor_total', 0):.2f}"
            
            # URL como link clicável
            url_produto = item.get('url', '')
            if url_produto:
                ws[f'E{row}'] = url_produto
                ws[f'E{row}'].hyperlink = Hyperlink(ref=f'E{row}', target=url_produto, tooltip="Clique para abrir o link")
                ws[f'E{row}'].font = Font(name='Arial', size=10, color='0000FF', underline='single')
            else:
                ws[f'E{row}'] = 'Sem URL'
            
            # Adicionar imagem do produto se existir
            produto_imagem = item.get('produto_imagem', '')
            if produto_imagem:
                try:
                    # Download da imagem
                    response = requests.get(produto_imagem, timeout=10)
                    if response.status_code == 200:
                        # Salvar imagem temporariamente
                        with tempfile.NamedTemporaryFile(delete=False, suffix='.jpg') as temp_file:
                            temp_file.write(response.content)
                            temp_file_path = temp_file.name
                        
                        # Redimensionar imagem para caber na célula
                        img = PILImage.open(temp_file_path)
                        img.thumbnail((100, 100), PILImage.Resampling.LANCZOS)
                        
                        # Salvar imagem redimensionada
                        img_path = temp_file_path.replace('.jpg', '_resized.jpg')
                        img.save(img_path, 'JPEG', quality=85)
                        
                        # Adicionar imagem ao Excel
                        excel_img = ExcelImage(img_path)
                        excel_img.width = 100
                        excel_img.height = 100
                        ws.add_image(excel_img, f'F{row}')
                        
                        # Limpar arquivo original, mas manter o redimensionado
                        os.unlink(temp_file_path)
                        
                        ws[f'F{row}'] = 'Imagem carregada'
                    else:
                        ws[f'F{row}'] = 'Erro ao carregar imagem'
                except Exception as e:
                    ws[f'F{row}'] = f'Erro: {str(e)[:20]}...'
            else:
                ws[f'F{row}'] = 'Sem imagem'
            
            # Aplicar formatação às células
            for col in range(1, 7):  # 6 colunas agora
                cell = ws.cell(row=row, column=col)
                cell.font = normal_font
                cell.border = thin_border
                if col in [2, 3, 4]:  # Quantidade, Valor Unit., Total
                    cell.alignment = center_alignment
                else:
                    cell.alignment = left_alignment
            
            # Ajustar altura da linha para acomodar a imagem
            ws.row_dimensions[row].height = 80
    
    # Resumo Financeiro
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "RESUMO FINANCEIRO"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].alignment = left_alignment
    ws.row_dimensions[row].height = 25
    
    # Dados do resumo financeiro
    moeda_symbol = 'R$' if ordem.get('moeda', 'BRL') == 'BRL' else '$' if ordem.get('moeda') == 'USD' else '¥'
    cotacao = ordem.get('cotacao_moeda', 1)
    valor_final = ordem.get('valor_final', 0)
    valor_brl = valor_final * cotacao
    
    # Calcular valores (usando dados da ordem se disponíveis)
    subtotal = ordem.get('subtotal', valor_final)
    comissao_agente = ordem.get('comissao_agente', 0)
    valor_transporte = ordem.get('valor_transporte', 0)
    taxas_adicionais = ordem.get('taxas_adicionais', 0)
    impostos_importacao = ordem.get('impostos_importacao', 0)
    total_ordem = ordem.get('valor_final', valor_final)
    
    resumo_data = [
        ['Subtotal:', f"{moeda_symbol} {subtotal:.2f}"],
        ['Comissão Agente:', f"{moeda_symbol} {comissao_agente:.2f}"],
        ['Valor Transporte:', f"{moeda_symbol} {valor_transporte:.2f}"],
        ['Taxas Adicionais:', f"{moeda_symbol} {taxas_adicionais:.2f}"],
        ['Impostos Importação:', f"{moeda_symbol} {impostos_importacao:.2f}"],
        ['Total:', f"{moeda_symbol} {total_ordem:.2f}"],
        ['Total em Reais:', f"R$ {valor_brl:.2f}"]
    ]
    
    for i, (label, value) in enumerate(resumo_data):
        row += 1
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = bold_font
        ws[f'A{row}'].fill = light_fill
        ws[f'A{row}'].border = thin_border
        
        ws[f'B{row}'] = value
        ws[f'B{row}'].font = normal_font
        ws[f'B{row}'].border = thin_border
        ws[f'B{row}'].alignment = right_alignment
        ws.merge_cells(f'B{row}:F{row}')
    
    # Observações
    if ordem.get('observacoes'):
        row += 2
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = "OBSERVAÇÕES"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws[f'A{row}'].alignment = left_alignment
        ws.row_dimensions[row].height = 25
        
        row += 1
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = ordem['observacoes']
        ws[f'A{row}'].font = normal_font
        ws[f'A{row}'].alignment = left_alignment
        ws[f'A{row}'].border = thin_border
    
    # Rodapé
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = f"Documento gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws[f'A{row}'].font = Font(name='Arial', size=8, italic=True)
    ws[f'A{row}'].alignment = center_alignment
    
    # Ajustar largura das colunas (6 colunas apenas)
    column_widths = [40, 12, 15, 15, 25, 20]  # Produto, Quantidade, Valor Unit., Total, URL, Imagem
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Salvar em buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Limpar arquivos temporários de imagens
    try:
        import glob
        temp_files = glob.glob('/tmp/tmp*_resized.jpg')
        for temp_file in temp_files:
            if os.path.exists(temp_file):
                os.unlink(temp_file)
    except Exception:
        pass  # Ignorar erros de limpeza
    
    # Retornar Excel
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=ordem_compra_{ordem_id}.xlsx"}
    )
